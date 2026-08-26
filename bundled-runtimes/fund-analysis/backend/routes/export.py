"""
数据导出路由 - 支持 Excel/CSV 格式导出基金列表、筛选结果、AI报告
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import io
import csv
import logging
from datetime import datetime

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["数据导出"])


def generate_excel(funds: list) -> io.BytesIO:
    """生成 Excel 文件流"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "基金列表"

        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = [
            "基金代码", "基金名称", "基金类型",
            "近1年收益", "近3年收益", "夏普比率", "最大回撤",
            "成立以来收益", "综合评分"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 数据行
        for row, fund in enumerate(funds, 2):
            ws.cell(row=row, column=1, value=fund.get("wind_code", ""))
            ws.cell(row=row, column=2, value=fund.get("name", ""))
            ws.cell(row=row, column=3, value=fund.get("type", ""))
            ws.cell(row=row, column=4, value=fund.get("return_1y"))
            ws.cell(row=row, column=5, value=fund.get("return_3y"))
            ws.cell(row=row, column=6, value=fund.get("sharpe_ratio"))
            ws.cell(row=row, column=7, value=fund.get("max_drawdown"))
            ws.cell(row=row, column=8, value=fund.get("return_since_inception"))
            ws.cell(row=row, column=9, value=fund.get("overall_score"))

        # 自动列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安装，请运行: pip install openpyxl")


def generate_csv(funds: list) -> io.BytesIO:
    """生成 CSV 文件流"""
    output = io.StringIO()
    writer = csv.writer(output)

    # 表头
    headers = [
        "基金代码", "基金名称", "基金类型",
        "近1年收益", "近3年收益", "夏普比率", "最大回撤",
        "成立以来收益", "综合评分"
    ]
    writer.writerow(headers)

    # 数据行
    for fund in funds:
        writer.writerow([
            fund.get("wind_code", ""),
            fund.get("name", ""),
            fund.get("type", ""),
            fund.get("return_1y"),
            fund.get("return_3y"),
            fund.get("sharpe_ratio"),
            fund.get("max_drawdown"),
            fund.get("return_since_inception"),
            fund.get("overall_score"),
        ])

    output.seek(0)
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))  # UTF-8 with BOM for Excel


@router.get("/funds")
def export_funds(
    format: str = Query("excel", description="导出格式: excel/csv"),
    fund_type: Optional[str] = Query(None, description="基金类型筛选"),
    keyword: Optional[str] = Query(None, description="搜索关键词"),
    limit: int = Query(1000, ge=1, le=5000, description="导出数量限制"),
    sort_by: str = Query("rank", description="排序字段"),
):
    """
    导出基金列表（Excel/CSV）

    限制：
    - 单次最多导出 5000 条记录
    - 指标一律读本地 metric_snapshots 面板，不实时调用外部数据源
      （历史版本逐基金调 Tushare 会打爆频率限额并阻塞事件循环，已废弃）
    """
    from service_registry import get_scoring_engine
    from repositories import get_fund_repo
    from database import get_engine
    from sqlalchemy import text

    # 限制导出数量
    if limit > 5000:
        raise HTTPException(status_code=400, detail="单次导出最多 5000 条记录")

    scoring_engine = get_scoring_engine()
    fund_repo = get_fund_repo()

    # 获取基金数据（仅本地库；不再降级外部数据源）
    db_result = fund_repo.list_funds(fund_type=fund_type, keyword=keyword, page=1, page_size=limit)
    db_funds = db_result.get("funds", []) if db_result.get("total", 0) > 0 else []

    # 批量读本地指标面板（每基金每指标取最新 as_of）
    metric_rows: dict = {}
    if db_funds:
        codes = [f.get("wind_code") for f in db_funds if f.get("wind_code")]
        engine = get_engine()
        with engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT DISTINCT ON (target_id, metric_name, metric_window)
                           target_id, metric_name, metric_window, metric_value
                    FROM metric_snapshots
                    WHERE target_type = 'fund'
                      AND target_id = ANY(:codes)
                      AND (
                        (metric_name = 'annualized_return' AND metric_window IN ('1y', '3y'))
                        OR metric_name IN ('sharpe_ratio', 'max_drawdown', 'annualized_volatility',
                                           'sortino_ratio', 'calmar_ratio', 'information_ratio', 'total_return')
                      )
                    ORDER BY target_id, metric_name, metric_window, as_of_date DESC
                    """
                ),
                {"codes": codes},
            ).fetchall()
        for target_id, metric_name, metric_window, metric_value in rows:
            metric_rows.setdefault(target_id, {})[(metric_name, metric_window)] = (
                float(metric_value) if metric_value is not None else None
            )

    def pick(code: str, metric_name: str, *windows: str):
        metrics = metric_rows.get(code, {})
        for window in windows or (None,):
            value = metrics.get((metric_name, window))
            if value is not None:
                return value
        return None

    funds = []
    for f in db_funds:
        wind_code = f.get("wind_code", "")
        if not wind_code:
            continue
        perf = {
            "annualized_return_1y": pick(wind_code, "annualized_return", "1y"),
            "annualized_return_3y": pick(wind_code, "annualized_return", "3y"),
            "total_return": pick(wind_code, "total_return"),
        }
        risk = {
            "max_drawdown": pick(wind_code, "max_drawdown", "1y", "3y"),
            "annualized_volatility_1y": pick(wind_code, "annualized_volatility", "1y", "3y"),
            "sharpe_ratio": pick(wind_code, "sharpe_ratio", "1y", "3y"),
            "sortino": pick(wind_code, "sortino_ratio", "1y", "3y"),
            "calmar_ratio": pick(wind_code, "calmar_ratio", "1y", "3y"),
            "information_ratio": pick(wind_code, "information_ratio", "1y", "3y"),
        }
        scoring = scoring_engine.score_fund(perf, risk, {})
        funds.append({
            "wind_code": wind_code,
            "name": f.get("name", ""),
            "type": f.get("type", ""),
            "return_1y": perf["annualized_return_1y"],
            "return_3y": perf["annualized_return_3y"],
            "sharpe_ratio": risk["sharpe_ratio"],
            "max_drawdown": risk["max_drawdown"],
            "return_since_inception": perf["total_return"],
            "overall_score": scoring.get("overall_score"),
        })

    # 生成文件
    filename = f"funds_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    if format.lower() == "csv":
        file_stream = generate_csv(funds)
        media_type = "text/csv; charset=utf-8-sig"
        filename += ".csv"
    else:
        file_stream = generate_excel(funds)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename += ".xlsx"

    return StreamingResponse(
        iter([file_stream.read()]),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
    )


@router.get("/screening-results")
async def export_screening_results(
    criteria_id: str = Query(..., description="筛选条件 ID"),
    format: str = Query("excel", description="导出格式: excel/csv"),
    limit: int = Query(1000, ge=1, le=5000),
):
    """
    导出筛选结果

    使用筛选条件 ID 获取之前保存的筛选结果
    """
    from repositories import get_screening_repo

    screening_repo = get_screening_repo()

    try:
        result = screening_repo.get_screening_result(criteria_id)
        if not result:
            raise HTTPException(status_code=404, detail="筛选结果不存在或已过期")

        funds = result.get("matched_funds", [])[:limit]

        filename = f"screening_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

        if format.lower() == "csv":
            file_stream = generate_csv(funds)
            media_type = "text/csv; charset=utf-8-sig"
            filename += ".csv"
        else:
            file_stream = generate_excel(funds)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename += ".xlsx"

        return StreamingResponse(
            iter([file_stream.read()]),
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"}
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export screening results failed: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/report/{report_id}")
async def export_report(
    report_id: str,
    format: str = Query("markdown", description="导出格式: markdown/pdf/html"),
):
    """
    导出基金研究报告

    支持格式：
    - markdown: Markdown 文本
    - html: 渲染后的 HTML
    - pdf: PDF 文件（需要安装 reportlab）
    """
    from repositories import get_report_repo

    report_repo = get_report_repo()

    try:
        report = report_repo.get_report(report_id)
        if not report:
            raise HTTPException(status_code=404, detail="报告不存在")

        content = report.get("content", "")
        title = report.get("title", "基金分析报告")

        if format.lower() == "markdown":
            return {
                "content": content,
                "filename": f"{title}.md"
            }
        elif format.lower() == "html":
            html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{title}</title>
    <style>
        body {{ font-family: "Microsoft YaHei", Arial, sans-serif; line-height: 1.6; max-width: 900px; margin: 0 auto; padding: 20px; }}
        h1 {{ color: #1a1a2e; border-bottom: 2px solid #4472C4; padding-bottom: 10px; }}
        h2 {{ color: #2d3748; margin-top: 30px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #e2e8f0; padding: 10px; text-align: left; }}
        th {{ background: #4472C4; color: white; }}
    </style>
</head>
<body>
{content}
</body>
</html>
            """
            return StreamingResponse(
                iter([html_content.encode("utf-8")]),
                media_type="text/html; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{title}.html"}
            )
        else:
            raise HTTPException(status_code=400, detail=f"不支持的格式: {format}")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export report failed: {e}")
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")
